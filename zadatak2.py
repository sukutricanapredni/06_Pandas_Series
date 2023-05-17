##1.Zadatak
##Ucitati excel file data for practice u python
##i iz sheet-a Solution kolone cost price, selling price i profit pretvoriti u serije
##svaka serija ce imati naziv kolone
import openpyxl as op
import pandas as panda

wb=op.load_workbook('Data-for-Practice.xlsx')
ws=wb['Solution']

cost=ws['E']
selling=ws['F']
s_cost=panda.Series([i.value for i in cost[4:92]])
s_cost.rename(cost[3].value,inplace=True)
##print(s_cost)

s_selling=panda.Series([i.value for i in selling[4:92]])
s_selling.rename(selling[3].value,inplace=True)
##print(s_selling)

profit=s_selling-s_cost
profit.rename('Profit',inplace=True)
##print(profit)

# print(cost)
# print(selling)
# for i in cost:
#     print(i.value)

##2.Zadatak
##Pronaci koji je najveci profit
##print(profit.max())

##3.Zadatak
##Pronaci ukupno nabavnu cenu (cost)
##print(s_cost.sum())

##4.Pronaci ukupnu prodajnu cenu (selling)
##print(s_selling.sum())

##5.Izracunati ukupan profit u procentima
##print((s_selling.sum()-s_cost.sum())/s_cost.sum()*100)

# l=[1,2,3,4,5]
# for i in l:
#     print(i)

for i in profit:
    if i>100:
        print(profit[profit==i])

